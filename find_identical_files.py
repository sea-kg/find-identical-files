#!/usr/bin/env python3

import os
import hashlib
import csv
import tkinter
from tkinter import filedialog
from tkinter import ttk
from openpyxl import Workbook
import threading
import subprocess
import platform


DIR_PATH = os.path.dirname(os.path.realpath(__file__))
DIR_PATH = os.path.normpath(DIR_PATH)

window = tkinter.Tk()
window.geometry('410x300')
window.title("Find Identical Files")

var_dirpath = tkinter.StringVar()
var_dirpath.set(DIR_PATH)

var_label2 = tkinter.StringVar()
var_label2.set("...")

var_label3 = tkinter.StringVar()
var_label3.set("...")

var_label4 = tkinter.StringVar()
var_label4.set("...")

label1 = tkinter.Label(text="Find Identical Files in Directory:")
label1.place(x=10, y=10)

entry1 = ttk.Entry(window, textvariable=var_dirpath)
entry1['state'] = 'readonly'
entry1.place(x=10, y=45, width=290)

def btn_click_select_dir():
    dirpath = filedialog.askdirectory(
        mustexist=False,
        initialdir=var_dirpath.get(),
    )
    if len(dirpath) > 0:
        var_dirpath.set(dirpath)

btn_select_dir = tkinter.Button(
    window,
    text='Select Dir',
    command=btn_click_select_dir,
)
btn_select_dir.place(x=310, y=40)

def get_all_files(_startdir):
    """ recursive find all files in dir """
    _ret = []
    _rec = [_startdir]
    while len(_rec) > 0:
        _dirpath = _rec[0]
        del _rec[0]
        for _file in os.listdir(_dirpath):
            _filepath = os.path.join(_dirpath, _file)
            if _file == '.git':
                continue
            if os.path.isdir(_filepath):
                _rec.append(_filepath)
                continue
            if os.path.isfile(_filepath):
                _ret.append(_filepath)
    return _ret

def start_find():
    try:
        _dirpath = var_dirpath.get()
        var_label2.set("...")
        var_label3.set("...")
        var_label4.set("...")
        progressbar['mode'] = 'indeterminate'
        progressbar.start(20)
        var_label2.set("Preapring a list of files...")
        _files = get_all_files(_dirpath)
        _lenfiles = len(_files)
        var_label2.set("All found files: " + str(_lenfiles))
        _filesizes = {}
        progressbar.stop()
        progressbar['mode'] = 'determinate'

        _idx = 0
        _lenfiles1 = 0
        _step_cur = 0
        progressbar.step(_step_cur)
        for _file in _files:
            _idx += 1
            _step = int((_idx / _lenfiles) * 100)
            if _step_cur != _step:
                # print(_step_cur, _idx)
                _step_cur = _step
                progressbar.step(_step_cur)
            # print(_file)
            _filesize = os.path.getsize(_file)
            if _filesize not in _filesizes:
                _filesizes[_filesize] = []
                _lenfiles1 += 1
            _filesizes[_filesize].append(_file)

        var_label3.set("Filtering by filesizes...")
        _idx = 0
        _filesizes_filtered = {}
        _lenfiles2 = 0
        _step_cur = 0
        progressbar.step(_step_cur)
        _found_files_after_first_filter = 0
        for _filesize in _filesizes:
            _idx += 1
            _step = int((_idx / _lenfiles1) * 100)
            if _step_cur != _step:
                _step_cur = _step
                progressbar.step(_step_cur)
            _len = len(_filesizes[_filesize])
            if _len > 1:
                _filesizes_filtered[_filesize] = _filesizes[_filesize]
                _found_files_after_first_filter += _len
                _lenfiles2 += 1
                # print(_filesize, _len)
        var_label3.set("Found files (after first filter by fsize): " + str(_found_files_after_first_filter))
        if _lenfiles2 == 0:
            var_label4.set("Not found")
        else:
            var_label4.set("Calculation hash for files...")
            _step_cur = 0
            progressbar.step(_step_cur)
            _idx = 0
            _lenfiles3 = 0
            _files_by_hashes = {}
            for _filesize in _filesizes_filtered:
                _idx += 1
                _step = int((_idx / _lenfiles2) * 100)
                if _step_cur != _step:
                    _step_cur = _step
                    progressbar.step(_step_cur)
                _len = len(_filesizes_filtered[_filesize])
                if _len > 1:
                    # print(_filesize, len(_filesizes_filtered[_filesize]))
                    for _file in _filesizes_filtered[_filesize]:
                        # print(_file)
                        with open(_file, 'rb', buffering=0) as _filer:
                            _hash = hashlib.md5(_filer.read()).hexdigest()
                            if _hash not in _files_by_hashes:
                                _files_by_hashes[_hash] = []
                                _lenfiles3 += 1
                            _files_by_hashes[_hash].append(_file)

            _filesizes_filtered2 = {}
            _found_files_after_second_filter = 0
            _idx = 0
            _lenfiles4 = 0
            _step_cur = 0
            progressbar.step(_step_cur)
            for _hash in _files_by_hashes:
                _idx += 1
                _step = int((_idx / _lenfiles3) * 100)
                if _step_cur != _step:
                    _step_cur = _step
                    progressbar.step(_step_cur)
                _len = len(_files_by_hashes[_hash])
                if _len > 1:
                    _found_files_after_second_filter += _len
                    _lenfiles4 += 1
                    _filesizes_filtered2[_hash] = _files_by_hashes[_hash]
            var_label4.set("Found files (after second filter by hash): " + str(_found_files_after_second_filter))

            if _lenfiles4 > 0:
                _filepath = "find-identical-files.xlsx"
                wb = Workbook()
                ws = wb.active  # grab the active worksheet
                ws['A1'] = _dirpath
                _idx = 2
                for _hash in _filesizes_filtered2:
                    _idx += 1
                    ws['A' + str(_idx)] = "filehash: " + _hash
                    for _file in _filesizes_filtered2[_hash]:
                        _idx += 1
                        ws['A' + str(_idx)] = _file[len(_dirpath)+1:]
                    _idx += 1
                wb.save(_filepath)
                if platform.system() == 'Windows':    # Windows
                    os.startfile(_filepath)
                else:                                   # linux variants
                    subprocess.call(('xdg-open', _filepath))

    except Exception as err:
        var_label2.set("ERROR")
        var_label3.set(str(err))
        var_label4.set("...")
        progressbar.stop()
        progressbar['mode'] = 'determinate'
    finally:
        btn_find.config(state=tkinter.ACTIVE)
        progressbar.stop()
        progressbar['mode'] = 'determinate'
        progressbar.step(0)

def btn_click_start_find():
    btn_find.config(state=tkinter.DISABLED)
    t = threading.Thread(target=start_find)
    t.start()

btn_find = tkinter.Button(
    window,
    text='Start Find Identical Files',
    command=btn_click_start_find,
)
btn_find.place(x=10, y=80)

label2 = tkinter.Label(textvariable=var_label2)
label2.place(x=10, y=120)

label3 = tkinter.Label(textvariable=var_label3)
label3.place(x=10, y=160)

label4 = tkinter.Label(textvariable=var_label4)
label4.place(x=10, y=200)

progressbar = ttk.Progressbar(window)  # , mode="indeterminate")
progressbar.place(x=10, y=240, width=380)

window.mainloop()
